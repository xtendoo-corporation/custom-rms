import base64
import html
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from urllib.parse import urlsplit

from odoo import _, api, fields, models
from odoo.osv import expression
from odoo.exceptions import AccessError, ValidationError

class IrAttachment(models.Model):
    _inherit = 'ir.attachment'

    def _default_knowledge_category_id(self):
        return self.env.ref(
            'rms_custom_knowledge.document_knowledge_category_general',
            raise_if_not_found=False,
        )

    def _get_default_upload_category(self):
        category_id = self.env.context.get('default_knowledge_category_id')
        if category_id:
            return self.env['document.knowledge.category'].browse(category_id).exists()
        return self._default_knowledge_category_id()

    is_knowledge_document = fields.Boolean(
        string='Knowledge Document',
        index=True,
        default=False,
    )
    knowledge_category_id = fields.Many2one(
        'document.knowledge.category',
        string='Directory',
        index=True,
        ondelete='restrict',
        default=_default_knowledge_category_id,
    )
    body_markdown = fields.Text(
        string='Markdown Content',
        readonly=True,
        copy=False,
    )
    body_html = fields.Html(
        string='Preview',
        compute='_compute_body_html',
        sanitize=False,
    )
    preview_url = fields.Char(
        string='Preview URL',
        compute='_compute_preview_url',
    )
    knowledge_preview_is_image = fields.Boolean(
        string='Image Preview',
        compute='_compute_knowledge_preview_card',
    )
    knowledge_file_icon_class = fields.Char(
        string='File Icon',
        compute='_compute_knowledge_preview_card',
    )
    last_upload_date_display = fields.Char(
        string='Fecha de ultima subida',
        compute='_compute_last_upload_date_display',
    )

    user_can_upload_here = fields.Boolean(
        string='Puede subir aqui',
        compute='_compute_user_can_upload_here',
    )
    is_saved_knowledge_document = fields.Boolean(
        string='Documento guardado',
        compute='_compute_is_saved_knowledge_document',
    )

    def _compute_is_saved_knowledge_document(self):
        for attachment in self:
            attachment.is_saved_knowledge_document = bool(attachment.id)

    @api.depends('write_date')
    def _compute_last_upload_date_display(self):
        for attachment in self:
            if not attachment.write_date:
                attachment.last_upload_date_display = False
                continue
            date_value = fields.Datetime.context_timestamp(attachment, attachment.write_date)
            attachment.last_upload_date_display = date_value.strftime('%d/%m/%Y %H:%M:%S')

    def _check_knowledge_manager_access(self):
        if not self.env.user.has_group('rms_custom_knowledge.group_knowledge_manager'):
            raise AccessError(_('Only Knowledge Managers can manage knowledge documents.'))

    @api.depends('knowledge_category_id')
    @api.depends_context('uid', 'default_knowledge_category_id')
    def _compute_user_can_upload_here(self):
        categories = self.mapped('knowledge_category_id')
        default_category = self._get_default_upload_category()
        if default_category:
            categories |= default_category
        uploadable_category_ids = categories._get_user_uploadable_category_ids(categories.ids) if categories else set()
        for attachment in self:
            category = attachment.knowledge_category_id or default_category
            attachment.user_can_upload_here = category.id in uploadable_category_ids if category else False

    @api.depends('datas', 'mimetype', 'name', 'is_knowledge_document', 'type', 'url')
    def _compute_body_html(self):
        for attachment in self:
            attachment.body_html = attachment._get_knowledge_preview_html()

    @api.depends('datas', 'mimetype', 'name', 'is_knowledge_document', 'type', 'url')
    def _compute_preview_url(self):
        for attachment in self:
            attachment.preview_url = attachment._get_knowledge_preview_url()

    @api.depends('mimetype', 'name', 'type')
    def _compute_knowledge_preview_card(self):
        for attachment in self:
            mimetype = attachment.mimetype or ''
            filename = (attachment.name or '').lower()
            attachment.knowledge_preview_is_image = (
                attachment.type != 'url' and mimetype.startswith('image/')
            )
            attachment.knowledge_file_icon_class = (
                'fa fa-link'
                if attachment.type == 'url'
                else attachment._get_knowledge_file_icon_class(mimetype, filename)
            )

    @api.constrains('type', 'url', 'is_knowledge_document', 'knowledge_category_id')
    def _check_knowledge_external_url(self):
        for attachment in self:
            if not attachment.is_knowledge_document or attachment.type != 'url':
                continue
            if not attachment._is_valid_knowledge_external_url():
                raise ValidationError(
                    _(
                        'Indica una URL web completa y válida que comience por '
                        'https:// o http://.'
                    )
                )

    def _is_valid_knowledge_external_url(self):
        self.ensure_one()
        if not self.url:
            return False
        url = self.url.strip()
        if any(character.isspace() for character in url):
            return False
        try:
            parsed_url = urlsplit(url)
            hostname = parsed_url.hostname
        except ValueError:
            return False
        return parsed_url.scheme.lower() in ('https', 'http') and bool(hostname)

    def action_open_knowledge_url(self):
        self.ensure_one()
        if self.type != 'url' or not self._is_valid_knowledge_external_url():
            raise ValidationError(_('Este recurso no contiene un enlace externo.'))
        return {
            'type': 'ir.actions.act_url',
            'url': self.url.strip(),
            'target': 'new',
        }

    def action_open_knowledge_preview(self):
        self.ensure_one()
        form_view = self.env.ref(
            'rms_custom_knowledge.view_attachment_knowledge_form_primary'
        )
        return {
            'type': 'ir.actions.act_window',
            'name': self.display_name,
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'views': [(form_view.id, 'form')],
            'target': 'current',
            'context': {
                **self.env.context,
                'rms_knowledge_mode': True,
            },
        }

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if isinstance(vals.get('url'), str):
                vals['url'] = vals['url'].strip()
            if self._values_create_knowledge_document(vals):
                self._check_knowledge_upload_access_for_values(vals)
                vals['is_knowledge_document'] = True
                vals['body_markdown'] = False
        return super().create(vals_list)

    @api.model
    def _get_user_readable_knowledge_category_ids(self):
        if (
            self.env.user.has_group('rms_custom_knowledge.group_knowledge_manager')
            or self.env.user.has_group('rms_custom_knowledge.group_knowledge_contributor')
        ):
            return self.env['document.knowledge.category'].search([]).ids
        return self.env['document.knowledge.category'].search([
            ('access_line_ids.user_id', '=', self.env.uid),
        ]).ids

    @api.model
    def search_panel_select_range(self, field_name, **kwargs):
        if (
            field_name == 'knowledge_category_id'
            and not self.env.user.has_group('rms_custom_knowledge.group_knowledge_manager')
            and not self.env.user.has_group('rms_custom_knowledge.group_knowledge_contributor')
        ):
            kwargs = dict(kwargs)
            kwargs['comodel_domain'] = [
                ('access_line_ids.user_id', '=', self.env.uid),
            ] + kwargs.get('comodel_domain', [])
        return super().search_panel_select_range(field_name, **kwargs)

    @api.model
    def _get_knowledge_view_domain(self):
        domain = [
            ('knowledge_category_id', '!=', False),
            ('name', 'not ilike', 'web.assets'),
            ('name', 'not ilike', 'bus.websocket'),
        ]
        if (
            self.env.user.has_group('rms_custom_knowledge.group_knowledge_manager')
            or self.env.user.has_group('rms_custom_knowledge.group_knowledge_contributor')
        ):
            return domain

        category_ids = self.env['document.knowledge.access'].sudo().search([
            ('user_id', '=', self.env.uid),
        ]).mapped('category_id').ids
        if not category_ids:
            return [('id', '=', 0)]
        return expression.AND([domain, [('knowledge_category_id', 'in', category_ids)]])

    @api.model
    def _domain_targets_knowledge_documents(self, domain):
        for item in domain or []:
            if isinstance(item, (list, tuple)):
                if item and isinstance(item[0], str) and item[0] in (
                    'knowledge_category_id',
                    'knowledge_category_id.access_line_ids.user_id',
                    'is_knowledge_document',
                ):
                    return True
                if self._domain_targets_knowledge_documents(item):
                    return True
        return False

    @api.model
    def _must_apply_knowledge_view_domain(self, domain):
        return bool(
            self.env.context.get('rms_knowledge_mode')
            or self.env.context.get('default_is_knowledge_document')
            or self._domain_targets_knowledge_documents(domain)
        )

    @api.model
    def _search(self, domain, offset=0, limit=None, order=None, *, active_test=True, bypass_access=False):
        if not bypass_access and self._must_apply_knowledge_view_domain(domain):
            domain = expression.AND([domain, self._get_knowledge_view_domain()])
            return super()._search(
                domain,
                offset=offset,
                limit=limit,
                order=order,
                active_test=active_test,
                bypass_access=True,
            )
        return super()._search(
            domain,
            offset=offset,
            limit=limit,
            order=order,
            active_test=active_test,
            bypass_access=bypass_access,
        )

    def _get_rms_knowledge_accessible_attachments(self, operation):
        knowledge_attachment_ids = set(self.sudo().filtered('knowledge_category_id').ids)
        if not knowledge_attachment_ids:
            return self.browse()

        knowledge_attachments = self.browse(knowledge_attachment_ids)
        if self.env.user.has_group('rms_custom_knowledge.group_knowledge_manager'):
            return knowledge_attachments
        if operation == 'unlink':
            return self.browse()
        if self.env.user.has_group('rms_custom_knowledge.group_knowledge_contributor'):
            return knowledge_attachments

        sudo_knowledge_attachments = knowledge_attachments.sudo()
        access_domain = [
            ('user_id', '=', self.env.uid),
            ('category_id', 'in', sudo_knowledge_attachments.mapped('knowledge_category_id').ids),
        ]
        if operation in ('write', 'create', 'unlink'):
            access_domain.append(('permission', '=', 'read_upload'))

        allowed_category_ids = set(
            self.env['document.knowledge.access'].sudo().search(access_domain).mapped('category_id').ids
        )
        allowed_attachment_ids = set(
            sudo_knowledge_attachments.filtered(
                lambda attachment: attachment.knowledge_category_id.id in allowed_category_ids
            ).ids
        )
        return self.browse(allowed_attachment_ids)

    def _check_access(self, operation):
        allowed_knowledge_attachments = self._get_rms_knowledge_accessible_attachments(operation)
        remaining = self - allowed_knowledge_attachments
        if not remaining:
            return None
        return super(IrAttachment, remaining)._check_access(operation)

    def write(self, vals):
        if isinstance(vals.get('url'), str):
            vals = dict(vals)
            vals['url'] = vals['url'].strip()
        if vals.get('knowledge_category_id') and not vals.get('is_knowledge_document'):
            vals = dict(vals)
            vals['is_knowledge_document'] = True

        knowledge_attachments = self.filtered('is_knowledge_document')
        if vals.get('is_knowledge_document') or vals.get('knowledge_category_id'):
            knowledge_attachments |= self

        if knowledge_attachments:
            for attachment in knowledge_attachments:
                attachment._check_knowledge_upload_access_for_write(vals)

        if vals.get('datas') and knowledge_attachments:
            vals = dict(vals)
            vals['body_markdown'] = False
            knowledge_attachments._unlink_html_preview_attachments()
        return super().write(vals)

    def unlink(self):
        if any(self.mapped('is_knowledge_document')):
            self._check_knowledge_manager_access()
            self._unlink_html_preview_attachments()
        return super().unlink()

    @api.model
    def _values_create_knowledge_document(self, vals):
        return bool(
            vals.get('is_knowledge_document')
            or self.env.context.get('default_is_knowledge_document')
            or vals.get('knowledge_category_id')
        )

    @api.model
    def _get_knowledge_category_from_values(self, vals):
        category_id = vals.get('knowledge_category_id') or self.env.context.get('default_knowledge_category_id')
        if category_id:
            return self.env['document.knowledge.category'].browse(category_id).exists()
        return self._default_knowledge_category_id()

    @api.model
    def _check_knowledge_upload_access_for_values(self, vals):
        category = self._get_knowledge_category_from_values(vals)
        if not category or not category._check_user_can_upload_here():
            raise AccessError(_('No tienes permisos para subir archivos en este directorio.'))

    def _check_knowledge_upload_access_for_write(self, vals):
        self.ensure_one()
        if self.env.user.has_group('rms_custom_knowledge.group_knowledge_manager'):
            return

        if 'is_knowledge_document' in vals and not vals.get('knowledge_category_id'):
            raise AccessError(_('No tienes permisos para cambiar el estado Knowledge de este archivo.'))

        categories = self.knowledge_category_id
        if vals.get('knowledge_category_id'):
            categories |= self.env['document.knowledge.category'].browse(vals['knowledge_category_id']).exists()

        if not categories or any(not category._check_user_can_upload_here() for category in categories):
            raise AccessError(_('No tienes permisos para modificar archivos en este directorio.'))

    def _get_knowledge_preview_url(self):
        self.ensure_one()
        if self.type == 'url':
            return self.url.strip() if self._is_valid_knowledge_external_url() else False
        if not self.datas or not self.id:
            return False

        mimetype = self.mimetype or ''
        filename = (self.name or '').lower()
        if (
            self._is_textual_preview_file(mimetype, filename)
            or self._is_excel_file(mimetype, filename)
            or self._is_word_file(mimetype, filename)
        ):
            return False
        if mimetype.startswith('image/'):
            return '/web/image/ir.attachment/%s/datas' % self.id
        return '/web/content/%s?download=false' % self.id

    @api.model
    def _get_knowledge_file_icon_class(self, mimetype, filename):
        if mimetype == 'application/pdf' or filename.endswith('.pdf'):
            return 'fa fa-file-pdf-o'
        if self._is_excel_file(mimetype, filename) or filename.endswith(('.ods', '.csv')):
            return 'fa fa-file-excel-o'
        if self._is_word_file(mimetype, filename):
            return 'fa fa-file-word-o'
        if mimetype.startswith('text/') or self._is_textual_preview_file(mimetype, filename):
            return 'fa fa-file-text-o'
        if mimetype.startswith('image/'):
            return 'fa fa-file-image-o'
        if mimetype.startswith('video/'):
            return 'fa fa-file-video-o'
        if mimetype.startswith('audio/'):
            return 'fa fa-file-audio-o'
        if filename.endswith(('.zip', '.rar', '.7z', '.tar', '.gz')):
            return 'fa fa-file-archive-o'
        return 'fa fa-file-o'

    def _get_knowledge_preview_html(self):
        self.ensure_one()
        if self.type == 'url':
            if not self._is_valid_knowledge_external_url():
                return self._preview_empty_html(
                    _('Añade una URL web válida para abrir el enlace.')
                )
            safe_url = html.escape(self.url.strip(), quote=True)
            safe_name = html.escape(self.name or _('Enlace externo'))
            return """<div class="o_rms_knowledge_external_link_preview">
                <div class="o_rms_knowledge_external_link_preview_icon">
                    <i class="fa fa-link"></i>
                </div>
                <h3>%s</h3>
                <p class="text-muted">%s</p>
                <a class="btn btn-primary" href="%s" target="_blank" rel="noopener noreferrer">
                    <i class="fa fa-external-link me-1"></i> Abrir enlace
                </a>
            </div>""" % (safe_name, safe_url, safe_url)
        if not self.datas:
            return self._preview_empty_html(_("Upload a file to preview it here."))
        if not self.id:
            return self._preview_empty_html(_("Save the document to generate the preview."))

        mimetype = self.mimetype or ""
        filename = (self.name or "").lower()
        title = html.escape(self.name or _("Document"))
        if self._is_excel_file(mimetype, filename):
            return self._wrap_preview_inline_html(self._get_excel_preview_body())
        if self._is_word_file(mimetype, filename):
            return self._wrap_preview_inline_html(self._get_word_preview_body())
        if self._is_textual_preview_file(mimetype, filename):
            return self._wrap_preview_inline_html(self._get_textual_preview_body())

        preview_url = self._get_knowledge_preview_url()
        if not preview_url:
            return self._preview_empty_html(_("No preview available."))
        return self._iframe_preview(preview_url, title)

    def _get_or_create_html_preview_attachment(self):
        self.ensure_one()
        preview = self.env['ir.attachment'].sudo().search([
            ('res_model', '=', 'ir.attachment'),
            ('res_id', '=', self.id),
            ('res_field', '=', 'rms_knowledge_preview'),
        ], limit=1)
        html_document = self._wrap_preview_html_document(self._get_textual_preview_body())
        values = {
            'name': '%s.preview.html' % (self.name or self.id),
            'type': 'binary',
            'datas': base64.b64encode(html_document.encode()).decode(),
            'mimetype': 'text/html',
            'res_model': 'ir.attachment',
            'res_id': self.id,
            'res_field': 'rms_knowledge_preview',
        }
        if preview:
            preview.write(values)
        else:
            preview = self.env['ir.attachment'].sudo().create(values)
        return preview

    def _unlink_html_preview_attachments(self):
        previews = self.env['ir.attachment'].sudo().search([
            ('res_model', '=', 'ir.attachment'),
            ('res_id', 'in', self.ids),
            ('res_field', '=', 'rms_knowledge_preview'),
        ])
        previews.unlink()

    def _get_textual_preview_body(self):
        self.ensure_one()
        mimetype = self.mimetype or ''
        filename = (self.name or '').lower()
        if self._is_markdown_file(mimetype, filename):
            return self._markdown_to_html(self._decode_datas_as_text())
        return '<pre style="white-space: pre-wrap; word-break: break-word;">%s</pre>' % html.escape(
            self._decode_datas_as_text()
        )

    def _get_excel_preview_body(self):
        self.ensure_one()
        mimetype = self.mimetype or ""
        filename = (self.name or "").lower()
        if not self._is_modern_excel_file(mimetype, filename):
            return self._preview_empty_html(_("Excel .xls preview is not supported. Please upload an .xlsx file."))

        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter, range_boundaries
        except ImportError:
            return self._preview_empty_html(_("Install openpyxl to preview Excel files."))

        raw_content = self._get_raw_content()
        if not raw_content:
            return self._preview_empty_html(_("No readable content found in this file."))

        try:
            workbook = load_workbook(io.BytesIO(raw_content), read_only=False, data_only=True)
        except Exception:
            return self._preview_empty_html(_("This Excel file could not be previewed."))

        try:
            sheet = workbook.active
            max_rows = min(sheet.max_row or 1, 120)
            max_cols = min(sheet.max_column or 1, 40)
            table_ranges = self._get_excel_table_ranges(sheet, range_boundaries)
            rows = []
            colgroup = []
            for col_idx in range(1, max_cols + 1):
                width = sheet.column_dimensions[get_column_letter(col_idx)].width or 10
                colgroup.append('<col style="width: %spx;"/>' % int(max(42, min(width * 8, 260))))

            for row_idx in range(1, max_rows + 1):
                row_cells = []
                row_height = sheet.row_dimensions[row_idx].height
                row_style = 'height: %spx;' % int(row_height * 1.35) if row_height else ''
                for col_idx in range(1, max_cols + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    table_style = self._get_excel_table_cell_style(row_idx, col_idx, table_ranges)
                    cell_style = self._get_excel_cell_style(cell, table_style)
                    row_cells.append(
                        '<td style="%s">%s</td>' % (
                            html.escape(cell_style, quote=True),
                            html.escape(self._format_excel_cell(cell.value)),
                        )
                    )
                rows.append('<tr style="%s">%s</tr>' % (html.escape(row_style, quote=True), ''.join(row_cells)))
        finally:
            workbook.close()

        if not rows:
            return self._preview_empty_html(_("No readable content found in this file."))

        note = _("Showing first %(rows)s rows and %(cols)s columns of sheet %(sheet)s.") % {
            "rows": max_rows,
            "cols": max_cols,
            "sheet": sheet.title,
        }
        return """<div class="o_rms_knowledge_excel_preview">
            <div class="text-muted" style="margin-bottom: 8px;">%s</div>
            <div style="overflow: auto; max-height: 72vh; background: white;">
                <table style="border-collapse: collapse; table-layout: fixed; width: max-content;">%s%s</table>
            </div>
        </div>""" % (html.escape(note), ''.join(colgroup), ''.join(rows))

    def _get_word_preview_body(self):
        self.ensure_one()
        mimetype = self.mimetype or ""
        filename = (self.name or "").lower()
        if not self._is_modern_word_file(mimetype, filename):
            return self._preview_empty_html(_("Word .doc preview is not supported. Please upload a .docx file."))

        raw_content = self._get_raw_content()
        if not raw_content:
            return self._preview_empty_html(_("No readable content found in this file."))

        try:
            with zipfile.ZipFile(io.BytesIO(raw_content)) as document_zip:
                document_xml = document_zip.read('word/document.xml')
                relationships = self._get_docx_relationships(document_zip)
        except (KeyError, zipfile.BadZipFile):
            return self._preview_empty_html(_("This Word file could not be previewed."))

        try:
            document = ET.fromstring(document_xml)
        except ET.ParseError:
            return self._preview_empty_html(_("This Word file could not be previewed."))

        body = document.find(self._docx_tag('body'))
        if body is None:
            return self._preview_empty_html(_("No readable content found in this file."))

        blocks = []
        for child in body:
            if child.tag == self._docx_tag('p'):
                paragraph = self._docx_paragraph_to_html(child, relationships)
                if paragraph:
                    blocks.append(paragraph)
            elif child.tag == self._docx_tag('tbl'):
                table = self._docx_table_to_html(child, relationships)
                if table:
                    blocks.append(table)
            if len(blocks) >= 300:
                blocks.append('<p class="text-muted">%s</p>' % html.escape(_("Preview limited to the first 300 blocks.")))
                break

        if not blocks:
            return self._preview_empty_html(_("No readable content found in this file."))

        return '<div class="o_rms_knowledge_word_preview">%s</div>' % ''.join(blocks)

    @api.model
    def _get_docx_relationships(self, document_zip):
        try:
            rels_xml = document_zip.read('word/_rels/document.xml.rels')
        except KeyError:
            return {}

        try:
            rels = ET.fromstring(rels_xml)
        except ET.ParseError:
            return {}

        relationships = {}
        for rel in rels:
            rel_id = rel.attrib.get('Id')
            target = rel.attrib.get('Target')
            if rel_id and target:
                relationships[rel_id] = target
        return relationships

    @api.model
    def _docx_paragraph_to_html(self, paragraph, relationships):
        content = self._docx_inline_content_to_html(paragraph, relationships)
        if not content:
            return ''

        style = self._docx_paragraph_style(paragraph)
        normalized_style = re.sub(r'[^a-z0-9]', '', style.lower())
        if normalized_style in ('title',):
            return '<h1>%s</h1>' % content
        if normalized_style.startswith('heading'):
            level_match = re.search(r'heading(\d+)', normalized_style)
            level = int(level_match.group(1)) if level_match else 2
            level = max(1, min(level, 6))
            return '<h%s>%s</h%s>' % (level, content, level)
        if paragraph.find('.//' + self._docx_tag('numPr')) is not None:
            return '<p style="margin-left: 24px;">&bull; %s</p>' % content
        return '<p>%s</p>' % content

    @api.model
    def _docx_table_to_html(self, table, relationships):
        rows = []
        for row in table.findall(self._docx_tag('tr')):
            cells = []
            for cell in row.findall(self._docx_tag('tc')):
                cell_blocks = []
                for paragraph in cell.findall(self._docx_tag('p')):
                    paragraph_html = self._docx_inline_content_to_html(paragraph, relationships)
                    if paragraph_html:
                        cell_blocks.append('<div>%s</div>' % paragraph_html)
                cells.append('<td style="border: 1px solid #d1d5db; padding: 6px 8px; vertical-align: top;">%s</td>' % ''.join(cell_blocks))
            if cells:
                rows.append('<tr>%s</tr>' % ''.join(cells))
        if not rows:
            return ''
        return '<div style="overflow: auto; margin: 12px 0;"><table style="border-collapse: collapse; width: 100%;">%s</table></div>' % ''.join(rows)

    @api.model
    def _docx_inline_content_to_html(self, element, relationships):
        fragments = []
        for child in element:
            if child.tag == self._docx_tag('r'):
                fragments.append(self._docx_run_to_html(child))
            elif child.tag == self._docx_tag('hyperlink'):
                link_content = self._docx_inline_content_to_html(child, relationships)
                rel_id = child.attrib.get(self._docx_rel_tag('id'))
                href = relationships.get(rel_id, '')
                if href.startswith(('http://', 'https://', 'mailto:')):
                    fragments.append('<a href="%s" target="_blank" rel="noopener noreferrer">%s</a>' % (html.escape(href, quote=True), link_content))
                else:
                    fragments.append(link_content)
        return ''.join(fragment for fragment in fragments if fragment)

    @api.model
    def _docx_run_to_html(self, run):
        pieces = []
        for child in run:
            if child.tag == self._docx_tag('t'):
                pieces.append(html.escape(child.text or ''))
            elif child.tag == self._docx_tag('tab'):
                pieces.append('&emsp;')
            elif child.tag == self._docx_tag('br'):
                pieces.append('<br/>')

        content = ''.join(pieces)
        if not content:
            return ''

        properties = run.find(self._docx_tag('rPr'))
        if properties is not None:
            if properties.find(self._docx_tag('b')) is not None:
                content = '<strong>%s</strong>' % content
            if properties.find(self._docx_tag('i')) is not None:
                content = '<em>%s</em>' % content
            if properties.find(self._docx_tag('u')) is not None:
                content = '<u>%s</u>' % content
        return content

    @api.model
    def _docx_paragraph_style(self, paragraph):
        style = paragraph.find('./%s/%s' % (self._docx_tag('pPr'), self._docx_tag('pStyle')))
        return style.attrib.get(self._docx_tag('val'), '') if style is not None else ''

    @api.model
    def _docx_tag(self, tag):
        return '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}%s' % tag

    @api.model
    def _docx_rel_tag(self, tag):
        return '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}%s' % tag

    @api.model
    def _get_excel_table_ranges(self, sheet, range_boundaries):
        table_ranges = []
        for table in sheet.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            table_ranges.append({
                'min_col': min_col,
                'min_row': min_row,
                'max_col': max_col,
                'max_row': max_row,
                'show_row_stripes': bool(table.tableStyleInfo and table.tableStyleInfo.showRowStripes),
            })
        return table_ranges

    @api.model
    def _get_excel_table_cell_style(self, row_idx, col_idx, table_ranges):
        for table_range in table_ranges:
            if not (
                table_range['min_row'] <= row_idx <= table_range['max_row']
                and table_range['min_col'] <= col_idx <= table_range['max_col']
            ):
                continue
            if row_idx == table_range['min_row']:
                return {
                    'background': '#0070c0',
                    'color': '#ffffff',
                    'bold': True,
                    'border': '#ffffff',
                }
            if table_range['show_row_stripes'] and (row_idx - table_range['min_row']) % 2 == 1:
                return {'background': '#cfe8f7', 'border': '#00a2e8'}
            return {'background': '#ffffff', 'border': '#00a2e8'}
        return {}

    @api.model
    def _get_excel_cell_style(self, cell, table_style=None):
        table_style = table_style or {}
        styles = [
            'border: 1px solid %s' % table_style.get('border', '#d9e2ec'),
            'padding: 3px 6px',
            'white-space: nowrap',
            'overflow: hidden',
            'text-overflow: ellipsis',
            'height: 22px',
            'font-size: 13px',
            'vertical-align: middle',
        ]
        background = table_style.get('background') or self._excel_color_to_css(cell.fill.fgColor)
        if background:
            styles.append('background-color: %s' % background)
        color = table_style.get('color') or self._excel_color_to_css(cell.font.color)
        if color:
            styles.append('color: %s' % color)
        if cell.font.bold or table_style.get('bold'):
            styles.append('font-weight: 700')
        if cell.font.italic:
            styles.append('font-style: italic')
        if cell.alignment.horizontal:
            styles.append('text-align: %s' % self._excel_horizontal_alignment(cell.alignment.horizontal))
        if cell.alignment.vertical:
            styles.append('vertical-align: %s' % self._excel_vertical_alignment(cell.alignment.vertical))
        return '; '.join(styles) + ';'

    @api.model
    def _excel_color_to_css(self, color):
        if not color or color.type != 'rgb' or not color.rgb:
            return False
        rgb = color.rgb[-6:]
        if rgb.upper() in ('000000', 'FFFFFF'):
            return False
        return '#%s' % rgb

    @api.model
    def _excel_horizontal_alignment(self, value):
        return {
            'center': 'center',
            'right': 'right',
            'fill': 'left',
            'justify': 'justify',
            'centerContinuous': 'center',
            'distributed': 'justify',
        }.get(value, 'left')

    @api.model
    def _excel_vertical_alignment(self, value):
        return {
            'top': 'top',
            'center': 'middle',
            'bottom': 'bottom',
            'justify': 'middle',
            'distributed': 'middle',
        }.get(value, 'middle')

    @api.model
    def _format_excel_cell(self, value):
        if value is None:
            return ""
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat(sep=" ")
            except TypeError:
                return value.isoformat()
        return str(value)

    @api.model
    def _wrap_preview_inline_html(self, body):
        if not body:
            return self._preview_empty_html(_("No readable content found in this file."))
        return """<div class="o_rms_knowledge_preview o_rms_knowledge_markdown_preview" style="padding: 24px; line-height: 1.55;">%s</div>""" % body

    @api.model
    def _wrap_preview_html_document(self, body):
        return """<!doctype html>
<html>
<head>
<meta charset="utf-8"/>
<style>
body { font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 24px; color: #111827; line-height: 1.55; }
a { color: #714b9f; }
pre { background: #f8fafc; border: 1px solid #e5e7eb; border-radius: 6px; padding: 16px; overflow: auto; }
code { background: #f3f4f6; border-radius: 4px; padding: 1px 4px; }
blockquote { border-left: 4px solid #d1d5db; margin-left: 0; padding-left: 16px; color: #4b5563; }
img { max-width: 100%%; height: auto; }
table { border-collapse: collapse; width: 100%%; }
th, td { border: 1px solid #d1d5db; padding: 6px 8px; }
</style>
</head>
<body>%s</body>
</html>""" % body

    @api.model
    def _preview_empty_html(self, message):
        return '<div class="text-muted">%s</div>' % html.escape(message)

    @api.model
    def _iframe_preview(self, url, title):
        return """<div class="o_rms_knowledge_preview o_rms_knowledge_preview_iframe">
            <iframe src="%s" title="%s" style="width: 100%%; min-height: 72vh; border: 0;"></iframe>
        </div>""" % (url, title)

    @api.model
    def _iframe_srcdoc_preview(self, html_document, title):
        return """<div class="o_rms_knowledge_preview o_rms_knowledge_preview_iframe">
            <iframe srcdoc="%s" title="%s" sandbox="allow-popups allow-popups-to-escape-sandbox" style="width: 100%%; min-height: 72vh; border: 0;"></iframe>
        </div>""" % (html.escape(html_document, quote=True), title)

    @api.model
    def _is_word_file(self, mimetype, filename):
        return mimetype in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/vnd.ms-word.document.macroEnabled.12",
            "application/msword",
        ) or filename.endswith((".docx", ".docm", ".dotx", ".dotm", ".doc"))

    @api.model
    def _is_modern_word_file(self, mimetype, filename):
        return mimetype in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/vnd.ms-word.document.macroEnabled.12",
        ) or filename.endswith((".docx", ".docm", ".dotx", ".dotm"))

    @api.model
    def _is_excel_file(self, mimetype, filename):
        return mimetype in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel.sheet.macroEnabled.12",
            "application/vnd.ms-excel",
        ) or filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"))

    @api.model
    def _is_modern_excel_file(self, mimetype, filename):
        return mimetype in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel.sheet.macroEnabled.12",
        ) or filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm"))

    @api.model
    def _is_markdown_file(self, mimetype, filename):
        return mimetype in ('text/markdown', 'text/x-markdown') or filename.endswith(('.md', '.markdown'))

    @api.model
    def _is_textual_preview_file(self, mimetype, filename):
        return (
            self._is_markdown_file(mimetype, filename)
            or mimetype.startswith('text/')
            or mimetype in ('application/json', 'application/xml')
            or filename.endswith(('.txt', '.csv', '.log', '.rst', '.json', '.xml', '.yaml', '.yml'))
        )

    def _get_raw_content(self):
        self.ensure_one()
        if self.store_fname:
            return self._file_read(self.store_fname)
        return self.db_datas or b""

    def _decode_datas_as_text(self):
        self.ensure_one()
        raw_content = self._get_raw_content()

        if isinstance(raw_content, str):
            raw_content = raw_content.encode()

        for encoding in ("utf-8", "latin-1"):
            try:
                return raw_content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return raw_content.decode("utf-8", errors="replace")

    @api.model
    def _markdown_to_html(self, markdown_text):
        if not markdown_text:
            return ''

        try:
            import markdown
        except ImportError:
            return self._basic_markdown_to_html(markdown_text)

        return markdown.markdown(
            markdown_text,
            extensions=['extra', 'sane_lists', 'nl2br'],
            output_format='html5',
        )

    @api.model
    def _basic_markdown_to_html(self, markdown_text):
        blocks = []
        paragraph = []
        list_items = []

        def flush_paragraph():
            if paragraph:
                text = ' '.join(paragraph).strip()
                blocks.append('<p>%s</p>' % self._inline_markdown_to_html(text))
                paragraph.clear()

        def flush_list():
            if list_items:
                blocks.append('<ul>%s</ul>' % ''.join('<li>%s</li>' % item for item in list_items))
                list_items.clear()

        for raw_line in markdown_text.splitlines():
            line = raw_line.strip()
            if not line:
                flush_paragraph()
                flush_list()
                continue

            heading = re.match(r'^(#{1,6})\s+(.+)$', line)
            if heading:
                flush_paragraph()
                flush_list()
                level = len(heading.group(1))
                blocks.append('<h%s>%s</h%s>' % (level, self._inline_markdown_to_html(heading.group(2)), level))
                continue

            bullet = re.match(r'^[-*+]\s+(.+)$', line)
            if bullet:
                flush_paragraph()
                list_items.append(self._inline_markdown_to_html(bullet.group(1)))
                continue

            flush_list()
            paragraph.append(line)

        flush_paragraph()
        flush_list()
        return '<div class="o_rms_knowledge_markdown">%s</div>' % ''.join(blocks)

    @api.model
    def _inline_markdown_to_html(self, text):
        escaped = html.escape(text)
        escaped = re.sub(r'`([^`]+)`', r'<code>\1</code>', escaped)
        escaped = re.sub(r'\*\*([^*]+)\*\*', r'<strong>\1</strong>', escaped)
        escaped = re.sub(r'__([^_]+)__', r'<strong>\1</strong>', escaped)
        escaped = re.sub(r'(?<!\*)\*([^*]+)\*(?!\*)', r'<em>\1</em>', escaped)
        escaped = re.sub(r'\[([^\]]+)\]\((https?://[^)]+)\)', r'<a href="\2" target="_blank" rel="noopener noreferrer">\1</a>', escaped)
        return escaped
